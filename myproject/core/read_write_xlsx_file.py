import openpyxl
from random import randint, random


def read_sheet(filename):
    '''
    Read sheet and return a list of values.
    '''
    wb = openpyxl.load_workbook(filename)
    ws = wb['Sheet1']

    max_row = ws.max_row
    max_col = ws.max_column

    items = []

    for row in ws.iter_cols(max_col=max_col, max_row=max_row):
        sub_items = []
        for i, col in enumerate(row, 1):
            if i > 1:
                sub_items.append(col.value)
        items.append(sub_items)

    return items


def write_sheet(filename):
    '''
    Write on sheet.
    '''
    wb = openpyxl.load_workbook(filename)
    ws = wb['Sheet1']

    max_row = ws.max_row

    for row in range(2, max_row + 1):
        ws.cell(row=row, column=3).value = round(randint(1, 10) * random(), 2)
        ws.cell(row=row, column=4).value = f'=B{row}*C{row}'

    # Total
    ws.cell(row=max_row + 1, column=3).value = 'Total'
    ws.cell(row=max_row + 1, column=4).value = f'=SUM(D2:D{max_row})'

    return wb
